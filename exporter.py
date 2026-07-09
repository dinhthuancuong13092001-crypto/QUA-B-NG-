import datetime
from typing import List
from models import Team, Match, TeamStanding
from scheduler import get_team_display_name

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def export_to_excel(filepath: str, teams: List[Team], matches: List[Match], standings_func) -> bool:
    """
    Xuất dữ liệu giải đấu ra file Excel sử dụng openpyxl.
    Định dạng màu sắc chuyên nghiệp (Header xanh đậm, chữ trắng, viền rõ ràng, tự chỉnh độ rộng).
    """
    if not OPENPYXL_AVAILABLE:
        print("Lỗi: Cần cài đặt thư viện 'openpyxl' để xuất file Excel. Chạy lệnh: pip install openpyxl")
        return False
        
    wb = openpyxl.Workbook()
    
    # Định nghĩa các Style cho Excel
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Segoe UI", size=14, bold=True, color="1B365D")
    data_font = Font(name="Segoe UI", size=10)
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # ==================== SHEET 1: DANH SÁCH ĐỘI BÓNG ====================
    ws1 = wb.active
    ws1.title = "Danh sách Đội"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.cell(row=1, column=1, value="DANH SÁCH ĐỘI BÓNG THAM GIA GIẢI ĐẤU").font = title_font
    ws1.row_dimensions[1].height = 30
    
    headers1 = ["STT", "Tên Đội bóng", "Bảng đấu"]
    ws1.row_dimensions[3].height = 25
    
    for col_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    row_ptr = 4
    for idx, t in enumerate(filter(lambda x: x.id != "BYE", teams), 1):
        ws1.row_dimensions[row_ptr].height = 20
        c1 = ws1.cell(row=row_ptr, column=1, value=idx)
        c2 = ws1.cell(row=row_ptr, column=2, value=t.name)
        c3 = ws1.cell(row=row_ptr, column=3, value=t.group or "Không chia bảng")
        
        for c in (c1, c2, c3):
            c.font = data_font
            c.border = thin_border
        c1.alignment = align_center
        c2.alignment = align_left
        c3.alignment = align_center
        row_ptr += 1
        
    auto_fit_columns(ws1)
    
    # ==================== SHEET 2: LỊCH THI ĐẤU & KẾT QUẢ ====================
    ws2 = wb.create_sheet("Lịch thi đấu & Kết quả")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.cell(row=1, column=1, value="LỊCH THI ĐẤU VÀ KẾT QUẢ CÁC TRẬN ĐẤU").font = title_font
    ws2.row_dimensions[1].height = 30
    
    headers2 = ["Vòng đấu", "Bảng", "Đội nhà", "Tỷ số nhà", "-", "Tỷ số khách", "Đội khách", "Thời gian", "Trọng tài", "Thẻ nhà", "Thẻ khách"]
    ws2.row_dimensions[3].height = 25
    
    for col_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    row_ptr = 4
    for m in matches:
        ws2.row_dimensions[row_ptr].height = 20
        home_name = get_team_display_name(m.home_team_id, teams, matches)
        away_name = get_team_display_name(m.away_team_id, teams, matches)
        
        score_home = m.home_score if m.played and not m.is_bye else ""
        score_away = m.away_score if m.played and not m.is_bye else ""
        vs_str = "REST" if m.is_bye else "VS"
        
        time_str = "Nghỉ vòng này" if m.is_bye else f"{m.date_str} {m.time_str}".strip()
        ref_str = m.referee if m.referee else "Chưa phân công"
        
        card_home_str = f"V:{m.home_yellow_cards} | Đ:{m.home_red_cards}" if m.played and not m.is_bye else "-"
        card_away_str = f"V:{m.away_yellow_cards} | Đ:{m.away_red_cards}" if m.played and not m.is_bye else "-"
        
        vals = [
            m.round_name, m.group or "Knock-out", home_name, score_home, 
            vs_str, score_away, away_name, time_str, ref_str, card_home_str, card_away_str
        ]
        
        for col_idx, val in enumerate(vals, 1):
            c = ws2.cell(row=row_ptr, column=col_idx, value=val)
            c.font = data_font
            c.border = thin_border
            if col_idx in (1, 2, 4, 5, 6, 8, 10, 11):
                c.alignment = align_center
            else:
                c.alignment = align_left
        row_ptr += 1
        
    auto_fit_columns(ws2)
    
    # ==================== SHEET 3: BẢNG XẾP HẠNG ====================
    ws3 = wb.create_sheet("Bảng xếp hạng tổng hợp")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.cell(row=1, column=1, value="BẢNG XẾP HẠNG TOÀN GIẢI").font = title_font
    ws3.row_dimensions[1].height = 30
    
    groups_list = sorted(list(set([t.group for t in teams if t.group])))
    
    row_ptr = 3
    if groups_list:
        # Nếu có chia bảng đấu
        for g_name in groups_list:
            ws3.cell(row=row_ptr, column=1, value=f"BẢNG XẾP HẠNG BẢNG {g_name}").font = Font(name="Segoe UI", size=12, bold=True, color="1B365D")
            ws3.row_dimensions[row_ptr].height = 25
            row_ptr += 1
            
            headers3 = ["Hạng", "Đội bóng", "Trận", "Thắng", "Hòa", "Thua", "BT", "BP", "HS", "Thẻ phạt", "Điểm phạt", "Điểm số"]
            ws3.row_dimensions[row_ptr].height = 25
            for col_idx, h in enumerate(headers3, 1):
                cell = ws3.cell(row=row_ptr, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                cell.border = thin_border
            row_ptr += 1
            
            g_standings = standings_func(teams, matches, g_name)
            for idx, std in enumerate(g_standings, 1):
                ws3.row_dimensions[row_ptr].height = 20
                vals = [
                    idx, std.team_name, std.played, std.won, std.drawn, std.lost,
                    std.goals_for, std.goals_against, std.goal_difference,
                    f"V:{std.yellow_cards} | Đ:{std.red_cards}", std.discipline_points, std.points
                ]
                for col_idx, val in enumerate(vals, 1):
                    c = ws3.cell(row=row_ptr, column=col_idx, value=val)
                    c.border = thin_border
                    if col_idx == 12:
                        c.font = bold_font
                    else:
                        c.font = data_font
                        
                    if col_idx in (1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12):
                        c.alignment = align_center
                    else:
                        c.alignment = align_left
                row_ptr += 1
            row_ptr += 2  # Tạo khoảng cách giữa các bảng
    else:
        # Không chia bảng
        headers3 = ["Hạng", "Đội bóng", "Trận", "Thắng", "Hòa", "Thua", "BT", "BP", "HS", "Thẻ phạt", "Điểm phạt", "Điểm số"]
        ws3.row_dimensions[row_ptr].height = 25
        for col_idx, h in enumerate(headers3, 1):
            cell = ws3.cell(row=row_ptr, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
        row_ptr += 1
        
        all_standings = standings_func(teams, matches, None)
        for idx, std in enumerate(all_standings, 1):
            ws3.row_dimensions[row_ptr].height = 20
            vals = [
                idx, std.team_name, std.played, std.won, std.drawn, std.lost,
                std.goals_for, std.goals_against, std.goal_difference,
                f"V:{std.yellow_cards} | Đ:{std.red_cards}", std.discipline_points, std.points
            ]
            for col_idx, val in enumerate(vals, 1):
                c = ws3.cell(row=row_ptr, column=col_idx, value=val)
                c.border = thin_border
                if col_idx == 12:
                    c.font = bold_font
                else:
                    c.font = data_font
                    
                if col_idx in (1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12):
                    c.alignment = align_center
                else:
                    c.alignment = align_left
            row_ptr += 1
            
    auto_fit_columns(ws3)
    
    wb.save(filepath)
    return True

def auto_fit_columns(ws):
    """Căn chỉnh động độ rộng của cột dựa trên nội dung."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                val_str = str(cell.value)
                # Tính độ dài thô sơ, hỗ trợ hiển thị tiếng Việt có dấu
                val_len = len(val_str.encode('utf-8')) // 2 + 3
                if val_len > max_len:
                    max_len = val_len
        ws.column_dimensions[col_letter].width = max(max_len, 10)
